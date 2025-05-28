from rest_framework import viewsets, status, generics
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import action
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.timezone import now
from datetime import timedelta, datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
import io

from .models import Parceiro, CanalVenda, Interacao
from .serializers import (
    ParceiroSerializer,
    CanalVendaSerializer,
    InteracaoSerializer,
    InteracaoPendentesSerializer,
)
User = get_user_model()

class ParceiroViewSet(viewsets.ModelViewSet):
    queryset = Parceiro.objects.all()
    serializer_class = ParceiroSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.tipo_user == 'ADMIN':
            return Parceiro.objects.all()
        elif user.tipo_user == 'GESTOR':
            return Parceiro.objects.filter(canal_venda__in=user.canais_venda.all())
        elif user.tipo_user == 'VENDEDOR':
            return Parceiro.objects.filter(consultor=user.id_vendedor)
        return Parceiro.objects.none()


class UploadParceirosView(viewsets.ViewSet):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def create(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file_obj)
        except Exception as e:
            return Response({'erro': f'Erro ao ler arquivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                for _, row in df.iterrows():
                    canal_nome = str(row.get('Canal de Venda')).strip()
                    canal, _ = CanalVenda.objects.get_or_create(nome=canal_nome)

                    parceiro_data = {
                        'codigo': str(row.get('Código')).strip(),
                        'parceiro': row.get('Parceiro', '').strip(),
                        'classificacao': row.get('Classificação', '').strip(),
                        'consultor': str(row.get('Consultor')).strip(),
                        'cidade': row.get('Cidade', '').strip(),
                        'uf': row.get('UF', '').strip(),
                        'unidade': row.get('Unidade', '').strip(),
                        'canal_venda': canal,
                        'janeiro': row.get('Janeiro', 0) or 0,
                        'fevereiro': row.get('Fevereiro', 0) or 0,
                        'marco': row.get('Março', 0) or 0,
                        'abril': row.get('Abril', 0) or 0,
                        'maio': row.get('Maio', 0) or 0,
                        'junho': row.get('Junho', 0) or 0,
                        'julho': row.get('Julho', 0) or 0,
                        'agosto': row.get('Agosto', 0) or 0,
                        'setembro': row.get('Setembro', 0) or 0,
                        'outubro': row.get('Outubro', 0) or 0,
                        'novembro': row.get('Novembro', 0) or 0,
                        'dezembro': row.get('Dezembro', 0) or 0,
                        'janeiro_2': row.get('Janeiro 2', 0) or 0,
                        'fevereiro_2': row.get('Fevereiro 2', 0) or 0,
                        'marco_2': row.get('Março 2', 0) or 0,
                    }

                    parceiro_obj, _ = Parceiro.objects.update_or_create(
                        codigo=parceiro_data['codigo'],
                        defaults=parceiro_data
                    )
                    parceiro_obj.save()

            return Response({'mensagem': 'Parceiros importados com sucesso'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        identificador = request.data.get('identificador')
        senha = request.data.get('senha')

        user = (
            User.objects.filter(username=identificador).first()
            or User.objects.filter(email=identificador).first()
            or User.objects.filter(id_vendedor=identificador).first()
        )

        if user and user.check_password(senha):
            return Response({
                "message": "Login realizado com sucesso",
                "usuario": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "tipo_user": user.tipo_user,
                    "id_vendedor": user.id_vendedor,
                    "canais_venda": [canal.nome for canal in user.canais_venda.all()],
                }
            })
        else:
            return Response({"erro": "Credenciais inválidas"}, status=status.HTTP_401_UNAUTHORIZED)


class CanalVendaListView(generics.ListAPIView):
    queryset = CanalVenda.objects.all()
    serializer_class = CanalVendaSerializer
    permission_classes = [IsAuthenticated]


class ParceiroCreateUpdateView(generics.CreateAPIView):
    queryset = Parceiro.objects.all()
    serializer_class = ParceiroSerializer
    permission_classes = [IsAuthenticated]


class CanalVendaViewSet(viewsets.ModelViewSet):
    queryset = CanalVenda.objects.all()
    serializer_class = CanalVendaSerializer


# === INTERAÇÕES ===

class InteracaoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar interações com parceiros.
    """
    serializer_class = InteracaoSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.tipo_user == 'ADMIN':
            return Interacao.objects.all()
        elif user.tipo_user == 'GESTOR':
            return Interacao.objects.filter(
                parceiro__canal_venda__in=user.canais_venda.all()
            )
        elif user.tipo_user == 'VENDEDOR':
            return Interacao.objects.filter(usuario=user)
        return Interacao.objects.none()
    
    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """
        Exporta as interações para um arquivo Excel.
        """
        # Obter parâmetros de filtro
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        
        # Converter strings para datas
        try:
            if data_inicio_str:
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            else:
                data_inicio = (now() - timedelta(days=30)).date()
                
            if data_fim_str:
                data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            else:
                data_fim = now().date()
        except ValueError:
            return Response(
                {"erro": "Formato de data inválido. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filtrar interações
        queryset = self.get_queryset().filter(
            data_interacao__date__gte=data_inicio,
            data_interacao__date__lte=data_fim
        ).order_by('-data_interacao')
        
        # Criar workbook e worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interações"
        
        # Definir cabeçalhos
        headers = [
            "ID", "Parceiro", "Unidade", "Classificação", "Status", 
            "Tipo de Interação", "Data da Interação", "Usuário", "Entrou em Contato"
        ]
        
        # Estilo para cabeçalho
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Adicionar cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Adicionar dados
        for row_num, interacao in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=interacao.id)
            ws.cell(row=row_num, column=2, value=interacao.parceiro.parceiro)
            ws.cell(row=row_num, column=3, value=interacao.parceiro.unidade)
            ws.cell(row=row_num, column=4, value=interacao.parceiro.classificacao)
            ws.cell(row=row_num, column=5, value=interacao.parceiro.status)
            ws.cell(row=row_num, column=6, value=interacao.get_tipo_display())
            ws.cell(row=row_num, column=7, value=interacao.data_interacao.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row_num, column=8, value=interacao.usuario.username)
            ws.cell(row=row_num, column=9, value="Sim" if interacao.entrou_em_contato else "Não")
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Salvar para buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Criar resposta HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=interacoes_{data_inicio}_a_{data_fim}.xlsx'
        
        return response


class InteracoesHojeView(generics.ListAPIView):
    serializer_class = InteracaoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        hoje = now().date()
        return Interacao.objects.filter(data_interacao__date=hoje, usuario=self.request.user)


class InteracoesPendentesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        usuario = request.user
        hoje = now().date()
        limite_data = hoje - timedelta(days=3)
        
        # Filtrar parceiros de acordo com o perfil do usuário
        if usuario.tipo_user == 'VENDEDOR':
            parceiros = Parceiro.objects.filter(consultor=usuario.id_vendedor)
        elif usuario.tipo_user == 'GESTOR':
            parceiros = Parceiro.objects.filter(canal_venda__in=usuario.canais_venda.all())
        else:  # ADMIN
            parceiros = Parceiro.objects.all()
        
        # Obter o tipo de lista solicitada (pendentes ou interagidos)
        tipo_lista = request.query_params.get('tipo', 'pendentes')
        
        # Listas para armazenar parceiros pendentes e interagidos
        parceiros_pendentes = []
        parceiros_interagidos = []
        
        # Processar cada parceiro
        for parceiro in parceiros:
            # Obter a última interação do parceiro
            ultima_interacao = parceiro.interacoes.order_by('-data_interacao').first()
            
            # Verificar se o parceiro foi interagido hoje
            interagido_hoje = ultima_interacao and ultima_interacao.data_interacao.date() == hoje
            
            # Verificar se o parceiro está no período de bloqueio (3 dias)
            em_periodo_bloqueio = (
                ultima_interacao and 
                ultima_interacao.entrou_em_contato and 
                ultima_interacao.data_interacao.date() > limite_data and
                ultima_interacao.data_interacao.date() < hoje
            )
            
            # Adicionar à lista apropriada
            if interagido_hoje:
                parceiros_interagidos.append({
                    'id': parceiro.id,
                    'parceiro': parceiro.parceiro,
                    'unidade': parceiro.unidade,
                    'classificacao': parceiro.classificacao,
                    'status': parceiro.status,
                    'tipo': ultima_interacao.tipo if ultima_interacao else '',
                    'data_interacao': ultima_interacao.data_interacao if ultima_interacao else '',
                    'entrou_em_contato': ultima_interacao.entrou_em_contato if ultima_interacao else False,
                })
            elif not em_periodo_bloqueio:
                parceiros_pendentes.append({
                    'id': parceiro.id,
                    'parceiro': parceiro.parceiro,
                    'unidade': parceiro.unidade,
                    'classificacao': parceiro.classificacao,
                    'status': parceiro.status,
                    'tipo': '',
                    'data_interacao': '',
                    'entrou_em_contato': False,
                })
        
        # Retornar a lista solicitada
        if tipo_lista == 'interagidos':
            return Response(parceiros_interagidos)
        else:  # pendentes (padrão)
            return Response(parceiros_pendentes)
    
    @action(detail=False, methods=['get'])
    def metas(self, request):
        """
        Retorna informações sobre as metas diárias de interação.
        """
        usuario = request.user
        hoje = now().date()
        
        # Obter o número de interações realizadas hoje
        interacoes_hoje = Interacao.objects.filter(
            usuario=usuario,
            data_interacao__date=hoje
        ).count()
        
        # Meta diária (poderia ser configurável por usuário ou perfil)
        meta_diaria = 10
        
        # Calcular progresso
        progresso = min(interacoes_hoje / meta_diaria, 1.0) if meta_diaria > 0 else 0
        meta_atingida = interacoes_hoje >= meta_diaria
        
        return Response({
            'interacoes_realizadas': interacoes_hoje,
            'meta_diaria': meta_diaria,
            'progresso': progresso,
            'meta_atingida': meta_atingida
        })


class HistoricoInteracoesView(generics.ListAPIView):
    serializer_class = InteracaoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        parceiro_id = self.request.query_params.get('parceiro_id')
        return Interacao.objects.filter(parceiro_id=parceiro_id).order_by('-data_interacao')


class RegistrarInteracaoView(generics.CreateAPIView):
    serializer_class = InteracaoSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)
